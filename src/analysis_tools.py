"""
Analysis tools for manufacturing data.
Contains pandas-based functions for calculating cycle times, bottlenecks, delays, and generating reports.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Any
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


def calculate_cycle_times(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate cycle time (in minutes) for each job.
    
    Args:
        df: Cleaned DataFrame with start_time and end_time columns
        
    Returns:
        DataFrame with added 'cycle_time_minutes' column, sorted by cycle time descending
    """
    df = df.copy()
    
    if "start_time" not in df.columns or "end_time" not in df.columns:
        raise ValueError("DataFrame must have 'start_time' and 'end_time' columns")
    
    # Calculate cycle time in minutes
    df["cycle_time_minutes"] = (df["end_time"] - df["start_time"]).dt.total_seconds() / 60
    
    # Sort by cycle time descending
    df = df.sort_values("cycle_time_minutes", ascending=False, na_position="last")
    
    return df.reset_index(drop=True)


def find_bottleneck_process(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Find the process step that is the bottleneck (longest average cycle time).
    
    Args:
        df: DataFrame with cycle_time_minutes and process_step columns
        
    Returns:
        Dictionary with bottleneck analysis:
        {
            "bottleneck_process": str,
            "avg_cycle_time": float (minutes),
            "total_time": float (minutes),
            "job_count": int,
            "reason": str
        }
    """
    if "cycle_time_minutes" not in df.columns:
        df = calculate_cycle_times(df)
    
    if "process_step" not in df.columns:
        raise ValueError("DataFrame must have 'process_step' column")
    
    # Group by process step and calculate statistics
    process_stats = df.groupby("process_step").agg({
        "cycle_time_minutes": ["mean", "sum", "count"]
    }).reset_index()
    
    process_stats.columns = ["process_step", "avg_cycle_time", "total_time", "job_count"]
    
    # Find process with highest average cycle time
    bottleneck_idx = process_stats["avg_cycle_time"].idxmax()
    bottleneck_row = process_stats.loc[bottleneck_idx]
    
    return {
        "bottleneck_process": bottleneck_row["process_step"],
        "avg_cycle_time": round(bottleneck_row["avg_cycle_time"], 2),
        "total_time": round(bottleneck_row["total_time"], 2),
        "job_count": int(bottleneck_row["job_count"]),
        "reason": f"Highest average cycle time: {bottleneck_row['avg_cycle_time']:.2f} minutes",
    }


def find_delayed_jobs(df: pd.DataFrame) -> pd.DataFrame:
    """
    Find jobs that are delayed (end_time > due_date) or have 'late' status.
    
    Args:
        df: DataFrame with end_time, due_date, and optional status columns
        
    Returns:
        DataFrame of delayed jobs with reason column
    """
    df = df.copy()
    delayed_jobs = []
    
    for idx, row in df.iterrows():
        reason = None
        
        # Check if overdue (end_time > due_date)
        if pd.notna(row.get("due_date")) and pd.notna(row.get("end_time")):
            if row["end_time"] > row["due_date"]:
                days_late = (row["end_time"] - row["due_date"]).days
                reason = f"Completed {days_late} days late"
        
        # Check status field for delay indicators
        if pd.notna(row.get("status")):
            status_lower = str(row["status"]).lower()
            if any(keyword in status_lower for keyword in ["late", "delayed", "overdue", "behind"]):
                reason = f"Status indicates delay: {row['status']}"
        
        if reason:
            row_copy = row.copy()
            row_copy["delay_reason"] = reason
            delayed_jobs.append(row_copy)
    
    if delayed_jobs:
        return pd.DataFrame(delayed_jobs).reset_index(drop=True)
    else:
        return pd.DataFrame()


def summarise_by_machine(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Summarize job metrics by machine.
    
    Args:
        df: DataFrame with machine and cycle_time_minutes columns
        
    Returns:
        Dictionary with machine utilization summary:
        {
            "top_machines": [{"name": str, "job_count": int, "avg_cycle_time": float}, ...],
            "underutilized": [{"name": str, "job_count": int}, ...],
            "total_machines": int
        }
    """
    if "cycle_time_minutes" not in df.columns:
        df = calculate_cycle_times(df)
    
    if "machine" not in df.columns:
        return {"top_machines": [], "underutilized": [], "total_machines": 0}
    
    # Group by machine
    machine_stats = df.groupby("machine").agg({
        "cycle_time_minutes": ["mean", "count"]
    }).reset_index()
    
    machine_stats.columns = ["machine", "avg_cycle_time", "job_count"]
    machine_stats = machine_stats.sort_values("avg_cycle_time", ascending=False)
    
    # Top machines (highest load)
    top_machines = machine_stats.head(5).to_dict("records")
    for machine in top_machines:
        machine["avg_cycle_time"] = round(machine["avg_cycle_time"], 2)
        machine["job_count"] = int(machine["job_count"])
    
    # Underutilized (lowest load)
    avg_jobs = machine_stats["job_count"].mean()
    underutilized = machine_stats[machine_stats["job_count"] < avg_jobs * 0.5].to_dict("records")
    for machine in underutilized:
        machine["job_count"] = int(machine["job_count"])
    
    return {
        "top_machines": top_machines,
        "underutilized": underutilized,
        "total_machines": len(machine_stats),
    }


def summarise_by_process(df: pd.DataFrame) -> Dict[str, List[Dict[str, Any]]]:
    """
    Summarize job metrics by process step.
    
    Args:
        df: DataFrame with process_step and cycle_time_minutes columns
        
    Returns:
        Dictionary with process ranking:
        {
            "by_cycle_time": [{"process": str, "avg_cycle_time": float, "job_count": int}, ...],
            "by_job_volume": [{"process": str, "job_count": int, "total_time": float}, ...]
        }
    """
    if "cycle_time_minutes" not in df.columns:
        df = calculate_cycle_times(df)
    
    if "process_step" not in df.columns:
        return {"by_cycle_time": [], "by_job_volume": []}
    
    # Group by process
    process_stats = df.groupby("process_step").agg({
        "cycle_time_minutes": ["mean", "sum", "count"]
    }).reset_index()
    
    process_stats.columns = ["process_step", "avg_cycle_time", "total_time", "job_count"]
    
    # By cycle time (descending)
    by_cycle_time = process_stats.sort_values("avg_cycle_time", ascending=False).to_dict("records")
    for p in by_cycle_time:
        p["avg_cycle_time"] = round(p["avg_cycle_time"], 2)
        p["job_count"] = int(p["job_count"])
    
    # By job volume (descending)
    by_volume = process_stats.sort_values("job_count", ascending=False).to_dict("records")
    for p in by_volume:
        p["total_time"] = round(p["total_time"], 2)
        p["job_count"] = int(p["job_count"])
    
    return {
        "by_cycle_time": by_cycle_time,
        "by_job_volume": by_volume,
    }


def estimate_wip_location(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Estimate where work-in-progress (WIP) is concentrated.
    Looks for process steps where many jobs have null or future end_times.
    
    Args:
        df: DataFrame with process_step and end_time columns
        
    Returns:
        Dictionary with WIP analysis:
        {
            "high_wip_locations": [{"process": str, "wip_count": int}, ...],
            "rationale": str
        }
    """
    if "process_step" not in df.columns or "end_time" not in df.columns:
        return {"high_wip_locations": [], "rationale": "Missing required columns"}
    
    # Find jobs that are likely still in progress (null end_time or future end_time)
    now = pd.Timestamp.now()
    wip_mask = df["end_time"].isna() | (df["end_time"] > now)
    wip_jobs = df[wip_mask]
    
    if len(wip_jobs) == 0:
        return {
            "high_wip_locations": [],
            "rationale": "No jobs in progress detected"
        }
    
    # Group by process step
    wip_by_process = wip_jobs["process_step"].value_counts().reset_index()
    wip_by_process.columns = ["process_step", "wip_count"]
    
    high_wip = wip_by_process.head(5).to_dict("records")
    for item in high_wip:
        item["wip_count"] = int(item["wip_count"])
    
    return {
        "high_wip_locations": high_wip,
        "rationale": f"Based on {len(wip_jobs)} jobs with null or future end_time",
    }


def generate_summary_statistics(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Generate high-level summary statistics of the manufacturing data.
    
    Args:
        df: DataFrame with manufacturing data
        
    Returns:
        Dictionary with summary statistics
    """
    if "cycle_time_minutes" not in df.columns:
        df = calculate_cycle_times(df)
    
    stats = {
        "total_jobs": len(df),
        "total_processing_time": round(df["cycle_time_minutes"].sum(), 2) if "cycle_time_minutes" in df.columns else 0,
        "avg_cycle_time": round(df["cycle_time_minutes"].mean(), 2) if "cycle_time_minutes" in df.columns else 0,
        "median_cycle_time": round(df["cycle_time_minutes"].median(), 2) if "cycle_time_minutes" in df.columns else 0,
        "min_cycle_time": round(df["cycle_time_minutes"].min(), 2) if "cycle_time_minutes" in df.columns else 0,
        "max_cycle_time": round(df["cycle_time_minutes"].max(), 2) if "cycle_time_minutes" in df.columns else 0,
    }
    
    # Status breakdown if available
    if "status" in df.columns:
        stats["job_count_by_status"] = df["status"].value_counts().to_dict()
    
    # Process breakdown if available
    if "process_step" in df.columns:
        stats["job_count_by_process"] = df["process_step"].value_counts().to_dict()
    
    return stats


def generate_excel_report(
    results: Dict[str, Any],
    output_path: str,
    df: pd.DataFrame,
) -> str:
    """
    Generate a comprehensive Excel report with analysis results.
    Creates multiple sheets with different analyses.
    
    Args:
        results: Dictionary of analysis results (from agent)
        output_path: Path where to save the Excel file
        df: Cleaned DataFrame for detailed export
        
    Returns:
        Path to the generated Excel file
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    # Write summary statistics
    ws["A1"] = "Manufacturing Analysis Report"
    ws["A1"].font = Font(bold=True, size=14)
    
    row = 3
    ws[f"A{row}"] = "Key Metrics"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    
    row += 1
    if isinstance(results.get("summary_statistics"), dict):
        for key, value in results["summary_statistics"].items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            ws[f"B{row}"] = value
            row += 1
    
    # Add bottleneck information
    row += 2
    ws[f"A{row}"] = "Bottleneck Analysis"
    ws[f"A{row}"].font = Font(bold=True, size=12)
    
    row += 1
    if isinstance(results.get("bottleneck_analysis"), dict):
        for key, value in results["bottleneck_analysis"].items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            ws[f"B{row}"] = value
            row += 1
    
    # Create detailed data sheet
    if len(df) > 0:
        ws_data = wb.create_sheet("Detailed Data")
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws_data.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Write data
        for r_idx, row_data in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row_data, 1):
                ws_data.cell(row=r_idx, column=c_idx).value = value
    
    # Auto-adjust column widths
    for ws in wb.sheetnames:
        worksheet = wb[ws]
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)
    
    return str(output_path)
